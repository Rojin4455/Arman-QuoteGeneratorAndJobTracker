"""
Management command to update contact custom field with quote link from Excel file.

Usage:
    python manage.py update_contact_quote_link --file /path/to/Book.xlsx
    python manage.py update_contact_quote_link --file /path/to/Book.xlsx --dry-run
"""

from django.core.management.base import BaseCommand, CommandError
from accounts.models import Contact, GHLAuthCredentials, GHLCustomField
import requests
import openpyxl


class Command(BaseCommand):
    help = 'Update contact custom field with quote link from Excel file'

    def add_arguments(self, parser):
        parser.add_argument(
            '--file',
            type=str,
            required=True,
            help='Path to Excel file containing Contact Id and Quote Link columns',
        )
        parser.add_argument(
            '--dry-run',
            action='store_true',
            help='Run in dry-run mode (no API calls will be made)',
        )

    def transform_quote_link(self, url):
        """
        Transform quote link from quotenew.theservicepilot.com to services.theservicepilot.com
        """
        if not url or not isinstance(url, str):
            return None
        
        # Replace quotenew with services
        transformed_url = url.replace('quotenew.theservicepilot.com', 'services.theservicepilot.com')
        return transformed_url

    def update_contact_custom_field(self, contact_id, field_value, credentials):
        """
        Update a contact's custom field via GHL API
        """
        # Get Quote Link custom field ID dynamically
        try:
            quote_link_field = GHLCustomField.objects.get(
                account=credentials,
                field_name='Quote Link',
                is_active=True
            )
            quote_link_field.refresh_from_db()
            
            # Validate that we have a real field ID (not a placeholder)
            if not quote_link_field.ghl_field_id or quote_link_field.ghl_field_id == 'ghl_field_id' or len(quote_link_field.ghl_field_id) < 5:
                return False, f"Invalid Quote Link field ID in database: '{quote_link_field.ghl_field_id}'"
            
            custom_field_id = quote_link_field.ghl_field_id
        except GHLCustomField.DoesNotExist:
            return False, "Quote Link custom field not found for this account. Please create it in the database."
        except Exception as e:
            return False, f"Error fetching Quote Link field: {str(e)}"
        
        url = f'https://services.leadconnectorhq.com/contacts/{contact_id}'
        headers = {
            'Authorization': f'Bearer {credentials.access_token}',
            'Content-Type': 'application/json',
            'Version': '2021-07-28',
            'Accept': 'application/json'
        }
        
        update_data = {
            "customFields": [
                {
                    "id": str(custom_field_id),
                    "field_value": field_value
                }
            ]
        }
        
        try:
            response = requests.put(url, headers=headers, json=update_data)
            if response.status_code in [200, 201]:
                return True, None
            else:
                return False, f"API returned status {response.status_code}: {response.text}"
        except Exception as e:
            return False, str(e)

    def handle(self, *args, **options):
        file_path = options['file']
        dry_run = options['dry_run']

        # Validate file exists
        try:
            workbook = openpyxl.load_workbook(file_path)
        except FileNotFoundError:
            raise CommandError(f'File not found: {file_path}')
        except Exception as e:
            raise CommandError(f'Error opening Excel file: {str(e)}')

        # Get the first worksheet
        worksheet = workbook.active

        # Find column indices
        header_row = None
        contact_id_col = None
        quote_link_col = None

        # Find header row (first row with data)
        for row_idx, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
            if any(cell for cell in row if cell):
                header_row = row_idx
                # Find column indices
                for col_idx, cell_value in enumerate(row, start=1):
                    if cell_value:
                        cell_str = str(cell_value).strip()
                        if cell_str.lower() == 'contact id':
                            contact_id_col = col_idx
                        elif cell_str.lower() == 'quote link':
                            quote_link_col = col_idx
                break

        if header_row is None:
            raise CommandError('Could not find header row in Excel file')

        if contact_id_col is None:
            raise CommandError('Could not find "Contact Id" column in Excel file')

        if quote_link_col is None:
            raise CommandError('Could not find "Quote Link" column in Excel file')

        self.stdout.write(
            self.style.SUCCESS(
                f'Found columns: Contact Id (col {contact_id_col}), Quote Link (col {quote_link_col})'
            )
        )

        # Process data rows
        success_count = 0
        error_count = 0
        skipped_count = 0

        # Cache for credentials by location_id
        credentials_cache = {}

        for row_idx, row in enumerate(worksheet.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
            # Skip empty rows
            if not any(cell for cell in row):
                continue

            contact_id = row[contact_id_col - 1] if contact_id_col <= len(row) else None
            quote_link = row[quote_link_col - 1] if quote_link_col <= len(row) else None

            # Skip if contact_id is missing
            if not contact_id:
                self.stdout.write(
                    self.style.WARNING(f'Row {row_idx}: Skipping - missing Contact Id')
                )
                skipped_count += 1
                continue

            contact_id = str(contact_id).strip()

            # Skip if quote_link is missing
            if not quote_link:
                self.stdout.write(
                    self.style.WARNING(f'Row {row_idx}: Skipping contact {contact_id} - missing Quote Link')
                )
                skipped_count += 1
                continue

            quote_link = str(quote_link).strip()

            # Transform the quote link
            transformed_link = self.transform_quote_link(quote_link)

            if not transformed_link:
                self.stdout.write(
                    self.style.WARNING(f'Row {row_idx}: Skipping contact {contact_id} - invalid Quote Link')
                )
                skipped_count += 1
                continue

            # Skip if URL is already transformed
            if transformed_link == quote_link:
                self.stdout.write(
                    self.style.NOTICE(f'Row {row_idx}: Contact {contact_id} - URL already uses services subdomain')
                )
                skipped_count += 1
                continue

            # Find contact in database to get location_id
            try:
                contact = Contact.objects.get(contact_id=contact_id)
            except Contact.DoesNotExist:
                self.stdout.write(
                    self.style.ERROR(f'Row {row_idx}: Contact {contact_id} not found in database')
                )
                error_count += 1
                continue
            except Contact.MultipleObjectsReturned:
                self.stdout.write(
                    self.style.ERROR(f'Row {row_idx}: Multiple contacts found with ID {contact_id}')
                )
                error_count += 1
                continue

            # Get credentials for this location
            location_id = contact.location_id
            if location_id not in credentials_cache:
                credentials = GHLAuthCredentials.objects.filter(location_id=location_id).first()
                if not credentials:
                    # Fallback to first credentials if location-specific not found
                    credentials = GHLAuthCredentials.objects.first()
                if not credentials:
                    self.stdout.write(
                        self.style.ERROR(f'Row {row_idx}: No credentials found for location {location_id}')
                    )
                    error_count += 1
                    continue
                credentials_cache[location_id] = credentials
            else:
                credentials = credentials_cache[location_id]

            if dry_run:
                self.stdout.write(
                    self.style.NOTICE(
                        f'[DRY RUN] Row {row_idx}: Would update contact {contact_id} ({contact.first_name} {contact.last_name}) '
                        f'with quote link: {transformed_link}'
                    )
                )
                success_count += 1
            else:
                # Update the contact via API
                success, error_msg = self.update_contact_custom_field(
                    contact_id, transformed_link, credentials
                )

                if success:
                    self.stdout.write(
                        self.style.SUCCESS(
                            f'Row {row_idx}: Successfully updated contact {contact_id} ({contact.first_name} {contact.last_name})'
                        )
                    )
                    success_count += 1
                else:
                    self.stdout.write(
                        self.style.ERROR(
                            f'Row {row_idx}: Failed to update contact {contact_id}: {error_msg}'
                        )
                    )
                    error_count += 1

        # Summary
        self.stdout.write('')
        self.stdout.write(self.style.SUCCESS('=' * 60))
        self.stdout.write(self.style.SUCCESS('Summary:'))
        self.stdout.write(self.style.SUCCESS(f'  Successfully processed: {success_count}'))
        self.stdout.write(self.style.ERROR(f'  Errors: {error_count}'))
        self.stdout.write(self.style.WARNING(f'  Skipped: {skipped_count}'))
        self.stdout.write(self.style.SUCCESS('=' * 60))

        if dry_run:
            self.stdout.write(self.style.NOTICE('\nThis was a dry run. No actual updates were made.'))

